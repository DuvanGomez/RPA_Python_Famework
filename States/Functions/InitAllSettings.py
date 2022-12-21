from openpyxl import Workbook, load_workbook

def initAllSettings(pathFile):
    dictTemp = dict()
    try:
        wb = load_workbook(pathFile)
    except :
        print(str.format("Error: {0}", "esto no leyo nada")) #falta revisar cómo clasificar o imprimir mejor los menjsaes y así poderlos ayudar a controlar
    else:
        for sheetName in wb.sheetnames:
            sheet = wb[sheetName]
            if str(sheetName) != "Assets":
                for x in range(2, sheet.max_row + 1 ):
                    if sheet.cell(row=x, column=1).value != None:
                        dictTemp[sheet.cell(row=x, column=1).value] = sheet.cell(row=x, column=2).value
            else:
                for x in range(2, sheet.max_row + 1):
                    if sheet.cell(row=x, column=1).value != None:
                        #acá va la función para obtener los valores correspondiente a cada asset
                        dictTemp[sheet.cell(row=x, column=1).value] = sheet.cell(row=x, column=2).value

    finally:
        wb.close()
    
    return dictTemp

def readSetupFiles(Config):
    dictTemp = dict()
    
    if "SetupFiles" in Config:
        try:
            wb = Workbook
            for x in str(Config["SetupFiles"]).split("|"):
                if x.replace("FilePath", "Source") == "Local":
                    wb = load_workbook(Config[x])
                    for sheetName in str(Config[x.replace("FilePath", "Sheets")]).split("|"):
                        dictTemp[sheetName] = wb[sheetName]

                else:
                    #código para lectura de spreadsheets/Database
                    dictTemp["Error"] = x
            
        except:
            print("Algo va mal con la lectura")
        finally:
            wb.close()


    return dictTemp

